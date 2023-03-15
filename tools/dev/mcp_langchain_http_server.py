from mcp.server.fastmcp import FastMCP
import os
import re
import csv
import math
from collections import Counter
from typing import Optional

mcp = FastMCP("LangChainMCP")

@mcp.tool()
def ping(message: str) -> str:
    return f"pong: {message}"

@mcp.tool()
def soma(a: float, b: float) -> float:
    return a + b

@mcp.tool()
def list_dir(path: str = ".") -> list:
    try:
        entries = []
        for name in os.listdir(path):
            full = os.path.join(path, name)
            kind = "DIR" if os.path.isdir(full) else "FILE"
            entries.append(f"{kind}:{name}")
        return entries
    except Exception as e:
        return [f"error:{e}"]

@mcp.tool()
def read_text_file(path: str, max_bytes: int = 65536, encoding: str = "utf-8") -> str:
    try:
        if not os.path.exists(path):
            return "error:not found"
        max_bytes = max(1, min(max_bytes, 2_000_000))
        with open(path, "rb") as f:
            data = f.read(max_bytes)
        return data.decode(encoding, errors="replace")
    except Exception as e:
        return f"error:{e}"

@mcp.tool()
def workspace_glob(base_path: str, pattern: str) -> list:
    import glob
    try:
        base_path = base_path or "."
        search_path = os.path.join(base_path, pattern)
        return glob.glob(search_path, recursive=True)
    except Exception as e:
        return [f"error:{e}"]

@mcp.tool()
def grep_text(base_path: str, regex: str, include_exts: str = "", max_matches: int = 50) -> list:
    """Busca por regex em arquivos de texto a partir de base_path.
    include_exts: lista separada por vírgula (ex.: ".py,.html,.md")
    """
    try:
        base_path = base_path or "."
        pattern = re.compile(regex, re.IGNORECASE)
        exts = [e.strip() for e in include_exts.split(",") if e.strip()] if include_exts else []
        results = []
        for root, _, files in os.walk(base_path):
            for fn in files:
                fpath = os.path.join(root, fn)
                if exts and not any(fn.endswith(e) for e in exts):
                    continue
                try:
                    with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                        for i, line in enumerate(f):
                            if pattern.search(line):
                                results.append({"file": fpath, "line": i + 1, "text": line.strip()})
                                if len(results) >= max_matches:
                                    return results
                except Exception:
                    continue
        return results
    except Exception as e:
        return [f"error:{e}"]

@mcp.tool()
def summarize_csv(path: str, delimiter: str = ",", encoding: str = "utf-8", max_rows: int = 100000) -> dict:
    """Resumo simples de CSV: contagem de linhas, colunas, nomes e estatísticas básicas numéricas."""
    try:
        if not os.path.exists(path):
            return {"error": "not found"}
        with open(path, "r", encoding=encoding, errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, None)
            if headers is None:
                return {"error": "no headers"}
            row_count = 0
            cols = len(headers)
            sums = [0.0] * cols
            counts = [0] * cols
            freq = [Counter() for _ in range(cols)]
            for row in reader:
                row_count += 1
                if row_count > max_rows:
                    break
                for idx, val in enumerate(row):
                    val = val.strip()
                    num = None
                    try:
                        num = float(val.replace(".", "").replace(",", ".")) if val else None
                    except Exception:
                        num = None
                    if num is not None and not math.isnan(num):
                        sums[idx] += num
                        counts[idx] += 1
                    else:
                        if val:
                            freq[idx][val] += 1
            numeric_stats = {}
            for i, h in enumerate(headers):
                if counts[i] > 0:
                    numeric_stats[h] = {"count": counts[i], "sum": sums[i], "mean": (sums[i] / counts[i])}
            top_values = {}
            for i, h in enumerate(headers):
                if freq[i]:
                    top_values[h] = freq[i].most_common(3)
            return {
                "rows": row_count,
                "columns": cols,
                "headers": headers,
                "numeric": numeric_stats,
                "top_values": top_values
            }
    except Exception as e:
        return {"error": str(e)}

@mcp.tool()
def read_excel_summary(path: str, sheet: str | None = None, max_rows: int = 10000) -> dict:
    """Resumo simples de Excel (requer openpyxl)."""
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        return {"error": f"openpyxl not installed: {e}"}
    try:
        if not os.path.exists(path):
            return {"error": "not found"}
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        ws = wb[sheet or sheet_names[0]]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(list(row))
            if i >= max_rows:
                break
        headers = rows[0] if rows else []
        row_count = max(0, len(rows) - 1) if headers else len(rows)
        return {
            "sheets": sheet_names,
            "sheet": ws.title,
            "headers": headers,
            "rows": row_count,
            "sample": rows[1:6] if headers else rows[:5]
        }
    except Exception as e:
        return {"error": str(e)}

@mcp.tool()
def duckdb_sql(path: Optional[str] = None, sql: str = "SELECT 1 AS one", table_name: str = "t", max_rows: int = 10000, fmt: str = "table") -> dict:
    """Executa SQL com DuckDB de forma local e gratuita.
    - Se `path` for fornecido (CSV ou Parquet), registra como view `table_name` e você pode consultar com `SELECT * FROM table_name ...`.
    - Caso não forneça `path`, o SQL pode usar funções como `read_csv_auto('...')` diretamente.
    Retorna até `max_rows` linhas.
    """
    try:
        import duckdb  # type: ignore
    except Exception as e:
        return {"error": f"duckdb not installed: {e}"}
    try:
        con = duckdb.connect()
        if path:
            if not os.path.exists(path):
                return {"error": "not found"}
            lower = path.lower()
            if lower.endswith(".csv"):
                con.execute(f"CREATE OR REPLACE VIEW {table_name} AS SELECT * FROM read_csv_auto(?)", [path])
            elif lower.endswith(".parquet"):
                con.execute(f"CREATE OR REPLACE VIEW {table_name} AS SELECT * FROM read_parquet(?)", [path])
            else:
                return {"error": "unsupported file type (use .csv or .parquet)"}
        cur = con.execute(sql)
        cols = [d[0] for d in cur.description] if cur.description else []
        rows = cur.fetchmany(max_rows)
        if fmt == "json":
            data = [dict(zip(cols, r)) for r in rows]
            return {"columns": cols, "rows": data}
        else:
            return {"columns": cols, "rows": rows}
    except Exception as e:
        return {"error": str(e)}

@mcp.tool()
def pandas_summary(path: str, encoding: str = "utf-8", delimiter: Optional[str] = None, max_rows: int = 200000) -> dict:
    """Resumo com Pandas (local e gratuito) para CSV.
    - Retorna shape, colunas, dtypes, amostra e describe numérico.
    """
    try:
        import pandas as pd  # type: ignore
    except Exception as e:
        return {"error": f"pandas not installed: {e}"}
    try:
        if not os.path.exists(path):
            return {"error": "not found"}
        read_kwargs = {"encoding": encoding}
        if delimiter is not None:
            read_kwargs["sep"] = delimiter
        df = pd.read_csv(path, **read_kwargs)
        if len(df) > max_rows:
            df = df.head(max_rows)
        dtypes = {c: str(t) for c, t in df.dtypes.items()}
        sample = df.head(5).to_dict(orient="records")
        describe = df.describe(include="all", percentiles=[0.25, 0.5, 0.75]).to_dict()
        return {
            "shape": list(df.shape),
            "columns": list(df.columns),
            "dtypes": dtypes,
            "sample": sample,
            "describe": describe,
        }
    except Exception as e:
        return {"error": str(e)}

# Create FastAPI app for MCP HTTP
app = mcp.streamable_http_app()

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("MCP_PORT", "8710"))
    host = os.getenv("MCP_HOST", "127.0.0.1")
    uvicorn.run(app, host=host, port=port)